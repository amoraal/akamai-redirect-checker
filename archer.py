#!/usr/bin/python3
"""
Usage:
    python3 archer.py -s <source> -d <destination> (optional)

Example:
    python ./archer.py \
    -s https://blog.akamai.com/ \
    -d https://www.akamai.com/blog?
"""

from urllib3.util import connection
import requests
import dns.resolver
from urllib.parse import urlparse, unquote
import argparse
from openpyxl import load_workbook

parser = argparse.ArgumentParser()
parser.add_argument(
    "-f",
    "--file",
    help="Provide a file with two columns in which the first has the source and the second has the destination URLs."
)
parser.add_argument(
    "-s",
    "--source",
    help="URL that needs to be redirected. Will print the destination with status code if provided without destination."
)
parser.add_argument(
    "-d",
    "--destination",
    help="Excepted destination URL. Will show if the redirect is successful by returning a success or fail message."
)
parser.add_argument(
    "-rs",
    "--relaxed-source",
    help="Relaxes the matching criteria by auto adding http:// to the source URL if it is missing.",
    action='store_true'
)
parser.add_argument(
    "-rd",
    "--relaxed-destination",
    help="Relaxes the matching criteria by auto adding https:// to the destination URL if it is missing.",
    action='store_true'
)
parser.add_argument(
    "-p",
    "--production",
    help="Routes DNS requests over the production network instead of staging.",
    action='store_true'
)
parser.add_argument(
    "-v",
    "--verbose",
    help="Displays the found config and redirect trace.",
    action='store_true'
)
args = parser.parse_args()

def check_redirect(source_url, target_url):

    url_port = source_url.rpartition('://')[0]
    if args.relaxed_source and not url_port:
        source_url = "http://"+source_url
        url_port = source_url.rpartition('://')[0]
    elif not args.relaxed_source and not url_port:
        print('[Error] Please provide full source URL with protocol or pass the --relaxed-source flag')
        quit()

    if target_url:
        dest_url_port = target_url.rpartition('://')[0]
        if args.relaxed_destination and not dest_url_port:
            target_url = "https://"+target_url
        elif not args.relaxed_destination and not dest_url_port:
            print('[Error] Please provide full destination URL with protocol or pass the --relaxed-destination flag')
            quit()

    redir_domain = urlparse(source_url).netloc

    # Get the first CNAME (which should be the akamai edgekey)
    try:
        akamai_edgekey_prod = str(dns.resolver.resolve(redir_domain, 'CNAME')[0])[:-1]
    except dns.resolver.NoAnswer:
        print('[Error] The domain', redir_domain, 'does not have any CNAME records or is not behind Akamai')
        quit()
    except dns.resolver.NoNameservers:
        print('[Error] The domain', redir_domain, 'did not resolve. Does the domain exist?')
        quit()
    akamai_edgekey_staging = akamai_edgekey_prod.replace("edgekey", "edgekey-staging")
    if args.verbose:
        if args.file:
            print("├─\033[0;30;47m Configuration \033[0m")
        else:
            print("┌─\033[0;30;47m Configuration \033[0m")
        if args.production:
            print("├─ Network                :\033[;1;37;40m Production \033[0m")
        else:
            print("├─ Network                :\033[;1;37;40m Staging \033[0m")
        print("├─ Source                 :\033[;1;37;40m", source_url, '\033[0m')
        print("├─ Destination            :\033[;1;37;40m", target_url, '\033[0m')
        print("├─ Domain                 :\033[;1;37;40m", redir_domain, '\033[0m')
        print("├─ Akamai EdgeKey         :\033[;1;37;40m", akamai_edgekey_prod, '\033[0m')
        if not args.production:
            print("├─ Akamai EdgeKey Staging :\033[;1;37;40m", akamai_edgekey_staging, '\033[0m')

    if not args.production:
        """ Define specific DNS

            The urllib3.connection module subclasses httplib.HTTPConnection under the same name,
            having replaced the .connect() method with one that calls self._new_conn. 
            In turn, this delegates to urllib3.util.connection.create_connection(). 
            From: https://stackoverflow.com/questions/22609385/python-requests-library-define-specific-dns
        """
        _orig_create_connection = connection.create_connection

        def patched_create_connection(address, *args, **kwargs):
            """Wrap urllib3's create_connection to resolve the name elsewhere"""
            host, port = address
            hostname = akamai_edgekey_staging

            return _orig_create_connection((hostname, port), *args, **kwargs)

        connection.create_connection = patched_create_connection

    response = requests.get(source_url)

    if args.verbose:
        print("│\n├─\033[0;30;47m Redirect Trace \033[0m")
        if response.history:
            for resp in response.history:
                print("├─[\033[;1;36;40m" + str(resp.status_code) + "\033[0m]─[\033[;1;37;40m" +
                      str(unquote(resp.url)) + "\033[0m]")  # unquote shows the encoded version of the url
            print("├─[\033[;1;36;40m" + str(response.status_code) + "\033[0m]─[\033[;1;37;40m" +
                  str(unquote(response.url)) + "\033[0m]")  # unquote shows the encoded version of the url
        else:
            print("├─ Request was not redirected")

        print("│\n├─\033[0;30;47m Result \033[0m")

    # If a target URL was provided, show if the redirect succeeded
    if target_url:
        if target_url == response.url or target_url == unquote(response.url):
            if args.verbose:
                print('└─[\033[32m' + "Success" + '\033[0m]' + " Redirect works!")
            else:
                print("[Success] Redirect works!")
        else:
            if args.verbose:
                print('└─[\033[31m' + "Failed" + '\033[0m]' + " Final destination does not match target URL!")
            else:
                print("[Failed] Final destination does not match target URL!")
    else:
        if args.verbose:
            print("└─[\033[;1;36;40m" + str(response.status_code) + "\033[0m]─[\033[;1;37;40m" +
                  str(unquote(response.url)) + "\033[0m]")
        else:
            print('[' + str(response.status_code) + ']', unquote(response.url))


if not args.source and not args.file:
    print("No source provided. Please supply a source with --file or --source (--help for help)")
    quit()

if args.file and args.destination or args.file and args.source:
    print("Do not use the --source or --destination flags in combination with a file (--help for help)")
    quit()

if args.source:
    source = args.source
    try:
        target = args.destination
    except IndexError:
        target = False

    check_redirect(source, target)

if args.file:
    file = args.file
    wb = load_workbook(filename=file)
    ws = wb.worksheets[0]  # or use wb['Sheet1'] to select specific sheet

    increment = 0
    for col1, col2 in zip(ws['A'], ws['B']):
        source = col1.value
        target = col2.value
        increment = increment + 1
        if args.verbose:
            print('┌─[' + str(increment) + ']')
        else:
            print('[' + str(increment) + ']', end="")
        check_redirect(source, target)
